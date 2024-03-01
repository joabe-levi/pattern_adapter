from typing import Tuple, List
from openpyxl import load_workbook


class SpreadsheetCatalogBase:
    spreadsheet: str

    def _load_sheet(self, spreadsheet: str) -> load_workbook:
        wb = load_workbook(f'PatternAdapter/{spreadsheet}')
        sheet = wb.active
        return sheet

    def get_class_import(self, line: Tuple[str]) -> dict:
        """
        here you implement the class that was created informing when methods will be created at runtime
        :param line:
        :return:
        """
        return {}

    def get_list_from_catalog(self) -> List[dict]:
        list_catalog = list()
        sheet = self._load_sheet(self.spreadsheet)
        """
        Today it is static with only 20 rows for performance optimization.
        """
        for line in list(sheet.iter_rows(min_row=7, values_only=True))[:20]:
            list_catalog.append(
                self.get_class_import(line)
            )
        return list_catalog

    def get_list_by_code(self, list_code: List[int]) -> List[dict]:
        partial_list_catalog = list()
        sheet = self._load_sheet(self.spreadsheet)
        """
        Today it is static with only 20 rows for performance optimization.
        """
        for line in list(sheet.iter_rows(min_row=7, values_only=True))[:20]:
            parse = self.get_class_import(line)
            if parse.get_pdm_code() in list_code:
                partial_list_catalog.append(
                    parse.get_adapted_line()
                )
        return partial_list_catalog
